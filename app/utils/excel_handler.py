import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Tuple
from datetime import datetime
from pathlib import Path
from app.config import EXPORTS_DIR, ESTUDIOS_DISPONIBLES

class ExcelHandler:
    """Manejador de importación y exportación de archivos Excel"""
    
    @staticmethod
    def export_to_excel(registros: List[Dict], filename: str = None) -> Path:
        """
        Exporta registros a un archivo Excel con formato
        
        Args:
            registros: Lista de diccionarios con los datos
            filename: Nombre del archivo (opcional)
            
        Returns:
            Path del archivo generado
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"registros_{timestamp}.xlsx"
        
        filepath = EXPORTS_DIR / filename
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        
        # Encabezados
        headers = ["ID", "Nombres", "Apellidos", "Email", "Estudio", "Fecha de Registro"]
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Agregar datos
        for registro in registros:
            fecha = registro.get('fecha_registro', '')
            if isinstance(fecha, datetime):
                fecha = fecha.strftime("%Y-%m-%d %H:%M:%S")
            
            ws.append([
                registro.get('id', ''),
                registro.get('nombres', ''),
                registro.get('apellidos', ''),
                registro.get('email', ''),
                registro.get('estudio', ''),
                fecha
            ])
        
        # Ajustar ancho de columnas
        column_widths = [10, 20, 20, 30, 25, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Guardar archivo
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def get_sheet_names(filepath: Path) -> List[str]:
        """
        Obtiene los nombres de todas las hojas del archivo Excel
        
        Args:
            filepath: Ruta del archivo Excel
            
        Returns:
            Lista con nombres de las hojas
        """
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            return []
    
    @staticmethod
    def import_from_excel_multiple_sheets(filepath: Path, sheet_names: List[str] = None) -> Dict[str, Tuple[List[Dict], List[str]]]:
        """
        Importa registros desde múltiples hojas de un archivo Excel
        
        Args:
            filepath: Ruta del archivo Excel
            sheet_names: Lista de nombres de hojas a importar (None = todas)
            
        Returns:
            Diccionario con {nombre_hoja: (registros_válidos, errores)}
        """
        try:
            # Obtener todas las hojas disponibles
            xl_file = pd.ExcelFile(filepath)
            available_sheets = xl_file.sheet_names
            
            # Si no se especifican hojas, importar todas
            if not sheet_names:
                sheet_names = available_sheets
            
            results = {}
            
            for sheet_name in sheet_names:
                if sheet_name not in available_sheets:
                    results[sheet_name] = ([], [f"La hoja '{sheet_name}' no existe en el archivo"])
                    continue
                
                # Procesar cada hoja
                registros, errores = ExcelHandler._process_sheet(filepath, sheet_name)
                results[sheet_name] = (registros, errores)
            
            return results
            
        except Exception as e:
            return {"error": ([], [f"Error al leer el archivo Excel: {str(e)}"])}
    
    @staticmethod
    def _process_sheet(filepath: Path, sheet_name: str) -> Tuple[List[Dict], List[str]]:
        """
        Procesa una hoja específica del Excel
        
        Args:
            filepath: Ruta del archivo Excel
            sheet_name: Nombre de la hoja
            
        Returns:
            Tupla con (registros_válidos, errores)
        """
        try:
            # Leer hoja específica
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            
            # Mapeo de columnas (flexible para diferentes formatos)
            column_mapping = {
                'nombres': ['nombres', 'nombre', 'first_name', 'firstname'],
                'apellidos': ['apellidos', 'apellido', 'last_name', 'lastname'],
                'email': ['email', 'correo', 'e-mail', 'mail'],
                'estudio': ['estudio', 'estudios', 'carrera', 'programa', 'course']
            }
            
            # Normalizar nombres de columnas
            df.columns = df.columns.str.strip().str.lower()
            
            # Mapear columnas
            mapped_columns = {}
            for key, possible_names in column_mapping.items():
                for col in df.columns:
                    if col in possible_names:
                        mapped_columns[key] = col
                        break
            
            # Verificar que existan las columnas necesarias
            required_fields = ['nombres', 'apellidos', 'email', 'estudio']
            missing_fields = [field for field in required_fields if field not in mapped_columns]
            
            if missing_fields:
                return [], [f"Hoja '{sheet_name}': Faltan columnas: {', '.join(missing_fields)}"]
            
            # Renombrar columnas
            df = df.rename(columns={v: k for k, v in mapped_columns.items()})
            
            # Filtrar solo las columnas necesarias
            df = df[required_fields]
            
            # Limpiar datos
            df = df.dropna(subset=required_fields)
            df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
            
            registros_validos = []
            errores = []
            
            for idx, row in df.iterrows():
                try:
                    # Validar estudio
                    estudio = str(row['estudio']).strip()
                    if estudio not in ESTUDIOS_DISPONIBLES:
                        errores.append(
                            f"Hoja '{sheet_name}', Fila {idx + 2}: Estudio '{estudio}' no válido. "
                            f"Debe ser: {', '.join(ESTUDIOS_DISPONIBLES)}"
                        )
                        continue
                    
                    # Validar email básico
                    email = str(row['email']).strip().lower()
                    if '@' not in email or '.' not in email:
                        errores.append(f"Hoja '{sheet_name}', Fila {idx + 2}: Email '{email}' no válido")
                        continue
                    
                    registro = {
                        'nombres': str(row['nombres']).strip(),
                        'apellidos': str(row['apellidos']).strip(),
                        'email': email,
                        'estudio': estudio
                    }
                    
                    registros_validos.append(registro)
                    
                except Exception as e:
                    errores.append(f"Hoja '{sheet_name}', Fila {idx + 2}: Error - {str(e)}")
            
            return registros_validos, errores
            
        except Exception as e:
            return [], [f"Error al procesar hoja '{sheet_name}': {str(e)}"]
    
    @staticmethod
    def import_from_excel(filepath: Path) -> Tuple[List[Dict], List[str]]:
        """
        Importa registros desde un archivo Excel (solo primera hoja - retrocompatibilidad)
        
        Args:
            filepath: Ruta del archivo Excel
            
        Returns:
            Tupla con (registros_válidos, errores)
        """
        return ExcelHandler._process_sheet(filepath, 0)
    
    @staticmethod
    def create_template() -> Path:
        """
        Crea un archivo Excel de plantilla para importación
        
        Returns:
            Path del archivo de plantilla
        """
        filename = "plantilla_registros.xlsx"
        filepath = EXPORTS_DIR / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Registros"
        
        # Encabezados
        headers = ["Nombres", "Apellidos", "Email", "Estudio"]
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="39a900", end_color="39a900", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Agregar ejemplos
        ejemplos = [
            ["Juan Carlos", "Pérez García", "juan.perez@example.com", "Técnico"],
            ["María José", "González López", "maria.gonzalez@example.com", "Tecnólogo"],
            ["Carlos Alberto", "Rodríguez Martínez", "carlos.rodriguez@example.com", "Especialización"]
        ]
        
        for ejemplo in ejemplos:
            ws.append(ejemplo)
        
        # Agregar hoja con estudios disponibles
        ws_estudios = wb.create_sheet("Estudios Disponibles")
        ws_estudios.append(["Estudios Válidos"])
        
        header_cell = ws_estudios['A1']
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for estudio in ESTUDIOS_DISPONIBLES:
            ws_estudios.append([estudio])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws_estudios.column_dimensions['A'].width = 30
        
        wb.save(filepath)
        return filepath